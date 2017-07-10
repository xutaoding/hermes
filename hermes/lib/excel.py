# -*- coding: utf-8 -*-

""" 
1. `xlrd` Package write data in '.xls' of ext excel file, which urls including less than equal 2 << 32.
2. `xlsxWriter` Package is a Python module for writing files in the Excel 2007+ XLSX file format, 
    write data in '.xlsx' of ext excel file, which urls including greater than a million.
3. `openpyxl` Package is a Python library to read/write Excel 2010 xlsx/xlsm files

Priority recommendation to use `openpyxl` or `xlsxWriter` Python Package
"""

import logging
import openpyxl
import xlsxwriter

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


class Workbook(object):
    """ The class is suitable to write to massive large data to excel, 
        but huge large data  will generate `MemoryError` error, so you must notice 3 item.
        
        Note that:
        1. version of `openpyxl` package  is 2.4.7
        2. single sheet could  accommodate 1048576 rows
        3. to write medium large data, parameter `write_only` must is True
    """
    def __init__(self, filename, sheet_name=None, sheet_index=None, sheet_headers=None):
        self.filename = filename
        self.sheet_index = sheet_index or 0
        self.sheet_headers = sheet_headers or []
        self.sheet_name = sheet_name and sheet_name or 'Sheet'

        self.max_rows = 0
        self.rows_per = 1048576
        self.sheet = None
        self.workbook = openpyxl.Workbook(write_only=True)

        self._add_sheet()

    def _add_sheet(self):
        self.sheet_name = self.sheet_name + str(self.sheet_index)

        # 1. `sheet_index is Noe`: append sheet to `Worksheet`
        # 2. `sheet_index` is integer: insert to location of `Worksheet`
        self.sheet = self.workbook.create_sheet(self.sheet_name, self.sheet_index)
        self.max_rows += self.sheet_headers and 1 or 0

    @staticmethod
    def _check_type(row_value):
        assert isinstance(row_value, (tuple, list))

    def write(self, row_value):
        self._check_type(row_value)

        if self.max_rows == self.rows_per:
            self.max_rows -= self.rows_per
            self.sheet_index += 1

        self.max_rows += 1
        self.sheet.append(row_value)

        logger.info('openpyxl info -> sheet index:{}, row: {}'.format(self.sheet_index, self.max_rows))

    def cell(self):
        pass

    def save(self):
        self.workbook.save(filename=self.filename)


class CapacityBook(object):
    """ This class can write huge large data to excel.Not only no like `xlwt` only write 65536 row 
        at most, nut also no like `openpyxl` just suitable to medium large data.
    
        Note that:
        1. version of `xlsxwriter` is 0.9.6
        2. single sheet could  accommodate 1048576 rows
        3. Data is chinese characters, require to convert unicode, otherwise fail
        4. huge large data will be a bit slow and recommendations to optimize
    """

    def __init__(self, save_path, sheet_name=None, headers=None):
        self.sheet_name = sheet_name and sheet_name or 'Sheet'
        self.sheet_headers = headers and headers or []
        self.workbook = xlsxwriter.Workbook(save_path, {'constant_memory': True})
        self.worksheet = None
        self.sheet_names = None
        self.sheet_index = 0
        self.rows_per = 1048576
        self.max_rows = 0

        self._add_sheet()

    def _add_sheet(self):
        self.worksheet = self.workbook.add_worksheet(self.sheet_name + str(self.sheet_index))

        for i, header in enumerate(self.sheet_headers):
            self.worksheet.write(0, i, header)

        self.max_rows = self.max_rows if not self.sheet_headers else self.max_rows + 1

    def write(self, values=None):
        if self.max_rows == self.rows_per:
            self.max_rows -= self.rows_per
            self.sheet_index += 1
            self._add_sheet()

        values = values if values is not None else []

        for col, value in enumerate(values):
            self.worksheet.write(self.max_rows, col, value)

            logger.info('xlsxwriter info -> sheet index:{}, rows: {}'.format(self.sheet_index, self.max_rows))

        self.max_rows = self.max_rows if not values else self.max_rows + 1

    def close(self):
        self.workbook.close()


if __name__ == '__main__':
    wb = openpyxl.Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([100, 200, 300])

    # Python types will automatically be converted
    import datetime

    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("D:\\tmp\\sample.xlsx")
